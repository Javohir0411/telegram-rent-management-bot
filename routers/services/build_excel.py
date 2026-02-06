from __future__ import annotations

import logging
from collections import defaultdict
from datetime import datetime, date
from typing import List
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from bot_strings.enum_str import PRODUCT_TYPE_LABEL, SIZE_LABEL


def build_excel(lang: str, rents: List["Rent"], start_date: date, end_date: date) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ижара ҳисобот"

    # =======================
    # STYLES
    # =======================
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    bold_font = Font(bold=True)

    header_fill = PatternFill("solid", fgColor="1F4E79")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # =======================
    # TITLE
    # =======================
    ws["A1"] = f"Ижара ҳисобот: {start_date} — {end_date}"
    ws.merge_cells("A1:Q1")
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 24

    ws["A2"] = f"Жами ёзувлар: {len(rents)}"
    ws.merge_cells("A2:Q2")
    ws["A2"].alignment = left

    # =======================
    # HEADERS
    # =======================
    headers = [
        "ID",
        "Мижоз",
        "Телефон рақам",
        "Паспорт",
        "Маҳсулот",
        "Ҳажми",
        "Миқдор",
        "Бошланиш санаси",
        "Тугаш санаси",
        "Ижарага берилган сана",
        "Кунлар",
        "Кунлик нарх",
        "Тўлов ҳолати",
        "Ижара ҳолати",
        "Маҳсулот нархи",
        "Етказиб бериш",
        "Умумий сумма",
    ]

    ws.append(headers)

    header_row = 3
    ws.row_dimensions[header_row].height = 22

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # =======================
    # DATA
    # =======================
    start_data_row = 4
    client_totals = defaultdict(float)

    logging.info(f"RENTS: {rents}")
    for row_i, r in enumerate(rents, start=start_data_row):
        renter = r.renter
        print(f"{row_i} RENTER: {renter}")
        product = r.product
        print(f"{row_i} PRODUCT: {product}")

        renter_name = renter.renter_fullname if renter else ""
        print(f"{row_i} RENTER NAME: {renter_name}")
        renter_phone = renter.renter_phone_number if renter else ""
        print(f"{row_i} RENTER PHONE: {renter_phone}")
        renter_passport = renter.renter_passport_info if renter else ""
        print(f"{row_i} RENTER PASS: {renter_passport}")

        product_type = str(PRODUCT_TYPE_LABEL[lang][product.product_type]) if product else ""
        print(f"{row_i} PRODUCT TYPE: {product_type}")
        product_size = str(SIZE_LABEL[lang][product.product_size]) if product and product.product_size else ""
        print(f"{row_i} PRODUCT SIZE: {product_size}")
        price_per_day = float(product.price_per_day or 0) if product else 0
        print(f"{row_i} PRICE PER DAY: {price_per_day}")
        status = str(r.status or None)
        print(f"{row_i} STATUS: {status}")
        rent_status = str(r.rent_status or None)
        print(f"{row_i} RENT STATUS: {rent_status}")

        qty = int(r.quantity or 0)
        print(f"{row_i} QUANTITY: {qty}")

        sd = r.start_date.date() if isinstance(r.start_date, datetime) else r.start_date
        print(f"{row_i} START DATE: {sd}")
        ed = r.end_date.date() if isinstance(r.end_date, datetime) else r.end_date
        print(f"{row_i} END DATE: {ed}")
        cd = r.created_at.date() if isinstance(r.created_at, datetime) else r.created_at
        print(f"{row_i} CREATED AT DATE: {cd}")

        if ed is None:
            days = 0
        else:
            days = (ed - sd).days + 1
            if days < 1:
                days = 1

        delivery_price = float(r.delivery_price or 0)
        print(f"{row_i} DELIVERY PRICE: {delivery_price}")
        print(f"{row_i} {r.renter.renter_fullname} - {product.product_type} PRICE: {r.product_price}")
        # Hisobot qayta hisoblamaydi:
        if ed is None:
            product_price = 0.0
            print(f"{row_i} END DATE IS NONE: {product_price}")
        else:
            product_price = float(r.product_price or 0)
            print(f"{row_i} PRODUCT PRICE: {product_price}")

        total = product_price + delivery_price
        print(f"{row_i} TOTAL: {total}")

        client_totals[renter_name] += total

        ed_cell = ed if ed is not None else ""
        ws.append([
            r.id,
            renter_name,
            renter_phone,
            renter_passport,
            product_type,
            product_size,
            qty,
            sd,
            ed_cell,
            cd,
            days,
            price_per_day,
            status,
            rent_status,
            product_price,
            delivery_price,
            total,
        ])

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row_i, column=col)
            c.border = border
            c.alignment = left if col in (2, 3, 4, 5) else center

    # =======================
    # FORMATS
    # =======================
    money_cols = [12, 15, 16, 17]
    for r in range(start_data_row, ws.max_row + 1):
        for c in money_cols:
            ws.cell(row=r, column=c).number_format = '#,##0.00'

    ws.column_dimensions["H"].number_format = "yyyy-mm-dd"
    ws.column_dimensions["I"].number_format = "yyyy-mm-dd"

    # =======================
    # TOTAL ROW
    # =======================
    total_row = ws.max_row + 2
    #
    # ws[f"M{total_row}"] = "Жами:"
    # ws[f"M{total_row}"].font = Font(bold=True)
    # ws[f"M{total_row}"].alignment = center
    #
    # data_end_row = ws.max_row  # <-- buni TOTAL ROW dan oldin ol
    # total_row = data_end_row + 2
    # ws[f"N{total_row}"] = f"=SUM(N{start_data_row}:N{data_end_row})"
    # ws[f"N{total_row}"].number_format = '#,##0.00'
    # ws[f"N{total_row}"].font = Font(bold=True)
    # ws[f"N{total_row}"].alignment = center

    # =======================
    # CLIENT TOTALS
    # =======================
    client_start = total_row + 3
    ws[f"B{client_start - 1}"] = "Мижоз бўйича жами"
    ws[f"B{client_start - 1}"].font = bold_font
    totally_sum = 0
    for i, (client, total_sum) in enumerate(client_totals.items(), start=client_start):
        logging.info(f"CLIENTS TOTALS: {client_totals}")
        logging.info(f"CLIENTS TOTALS IIII: {i}")
        ws[f"B{i}"] = client
        ws[f"D{i}"] = total_sum
        ws[f"D{i}"].number_format = '#,##0.00'
        # ws[f"D{i}"].font = bold_font
        totally_sum += total_sum

    ws[f"B{i + 2}"].font = bold_font
    ws[f"B{i + 2}"] = "Жами:"
    ws[f"D{i + 2}"].font = bold_font
    ws[f"D{i + 2}"].number_format = '#,##0.00'
    ws[f"D{i + 2}"] = totally_sum

    # =======================
    # COLUMN WIDTHS
    # =======================
    widths = {
        1: 8, 2: 22, 3: 16, 4: 16, 5: 14,
        6: 10, 7: 8, 8: 14, 9: 14, 10: 8,
        11: 12, 12: 14, 13: 14, 14: 14, 15: 16
    }

    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
